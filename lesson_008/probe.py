import openpyxl

wb = openpyxl.load_workbook('C:/example.xlsx')
# print(type(wb))
print(wb.sheetnames)
sheet = wb['Sheet1']
print(sheet)
another_sheet = wb.active
name = another_sheet.title
print(name)
my_cell = sheet['B1']
print(my_cell.value)
print(another_sheet['B1'].value)
my_another_cell = wb.active.cell(row=1, column=2).value
print(my_another_cell)

for i in range(1, 8, 2):
    print(i, sheet.cell(row=i, column=2).value)

print(wb.active.max_row)

for rowOfCellObjects in sheet['A1':'C3']:
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate, cellObj.value)
    print('---Конец строки---')